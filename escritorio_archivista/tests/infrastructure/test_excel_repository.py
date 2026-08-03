"""Tests para el repositorio Excel y el offset dinámico de datos."""
import sys
import os
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

import pytest
from openpyxl import Workbook

from src.infrastructure.excel_repository import ExcelRepository

HEADER_ROWS = 2


def _crear_inventario(path: str, data_start_row: int, rows: int = 3):
    """Crea un .xlsx con cabecera del documento, encabezados de columna y datos."""
    wb = Workbook()
    ws = wb.active

    nrows = data_start_row - HEADER_ROWS - 1
    for i in range(nrows):
        ws.cell(row=i + 1, column=1, value=f"Cabecera {i + 1}")

    for h in range(HEADER_ROWS):
        r = data_start_row - HEADER_ROWS + h
        ws.cell(row=r, column=1, value="N° de Registro")
        ws.cell(row=r, column=2, value="Escribano")
        ws.cell(row=r, column=3, value="N° de Folios")
        ws.cell(row=r, column=4, value="N° de Prot")

    datos = [
        ("101", "García", "001r-002v", "1"),
        ("102", "López", "003r-004v", "2"),
        ("103", "Pérez", "005r", "3"),
    ]
    for i, (reg, esc, fol, prot) in enumerate(datos):
        r = data_start_row + i
        ws.cell(row=r, column=1, value=reg)
        ws.cell(row=r, column=2, value=esc)
        ws.cell(row=r, column=3, value=fol)
        ws.cell(row=r, column=4, value=prot)

    wb.save(path)


def _crear_inventario_con_ejemplos(path: str):
    """Crea un .xlsx con bloque de ejemplo y anotaciones antes de los datos reales."""
    wb = Workbook()
    ws = wb.active

    for i in range(9):
        ws.cell(row=i + 1, column=1, value=f"Cabecera {i + 1}")

    for h in range(HEADER_ROWS):
        r = 10 + h
        ws.cell(row=r, column=1, value="N° de Registro")
        ws.cell(row=r, column=2, value="Escribano")
        ws.cell(row=r, column=3, value="N° de Folios")
        ws.cell(row=r, column=4, value="N° de Prot")

    # Bloque de ejemplo (filas 12-15): registro de ejemplo + anotaciones
    ws.cell(row=12, column=1, value="1")
    ws.cell(row=12, column=2, value="DIEGO DE AGUILAR")
    ws.cell(row=12, column=3, value="981v-991v")
    ws.cell(row=12, column=4, value="36")
    ws.cell(row=13, column=2, value="(NOMBRES Y APELLIDOS - COMPLETO)")
    ws.cell(row=13, column=3, value="r= recto v=verso (SIN ESPACIO)")
    ws.cell(row=14, column=1, value="1")
    ws.cell(row=14, column=2, value="DIEGO DE AGUILAR")
    ws.cell(row=14, column=3, value="981v-991v")
    ws.cell(row=14, column=4, value="36")
    ws.cell(row=15, column=2, value="(NOMBRES Y APELLIDOS - COMPLETO)")
    ws.cell(row=15, column=3, value="r= recto v=verso ( - SIN ESPACIO)")

    # Sub-encabezados (filas 17-18)
    ws.cell(row=17, column=1, value="Protocolo N°1")
    ws.cell(row=18, column=1, value="Registro N°1")

    # Datos reales (filas 19+)
    datos = [
        ("101", "García", "001r-002v", "1"),
        ("102", "López", "003r-004v", "1"),
        ("103", "Pérez", "005r", "1"),
    ]
    for i, (reg, esc, fol, prot) in enumerate(datos):
        r = 19 + i
        ws.cell(row=r, column=1, value=reg)
        ws.cell(row=r, column=2, value=esc)
        ws.cell(row=r, column=3, value=fol)
        ws.cell(row=r, column=4, value=prot)

    wb.save(path)


class TestExcelRepository:
    """Tests del repositorio Excel con inicio de datos configurable."""

    def setup_method(self):
        self.repo = ExcelRepository()
        self.tmp_dir = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp_dir, "inventario.xlsx")

    def test_offset_default_empieza_en_fila_10(self):
        _crear_inventario(self.path, data_start_row=10)
        records = self.repo.cargar_registros(
            self.path, fila_datos_inicio=10, fila_inicio=10, fila_fin=500
        )
        assert [r.fila for r in records] == [10, 11, 12]

    def test_offset_configurable_empieza_en_fila_15(self):
        _crear_inventario(self.path, data_start_row=15)
        records = self.repo.cargar_registros(
            self.path, fila_datos_inicio=15, fila_inicio=10, fila_fin=500
        )
        assert [r.fila for r in records] == [15, 16, 17]

    def test_filtro_rango_de_filas(self):
        _crear_inventario(self.path, data_start_row=10)
        records = self.repo.cargar_registros(
            self.path, fila_datos_inicio=10, fila_inicio=11, fila_fin=11
        )
        assert [r.fila for r in records] == [11]

    def test_datos_correctos_por_registro(self):
        _crear_inventario(self.path, data_start_row=10)
        records = self.repo.cargar_registros(
            self.path, fila_datos_inicio=10, fila_inicio=10, fila_fin=500
        )
        assert records[0].registro == "101"
        assert records[0].escribano == "García"
        assert records[0].folios == "001r-002v"
        assert records[0].protocolo == "1"

    def test_extraer_metadatos_detecta_siglo(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Inventario")
        ws.cell(row=2, column=1, value="Sección: XIX")
        ws.cell(row=3, column=1, value="Código del fondo: N07")
        for h in range(HEADER_ROWS):
            r = 10 - HEADER_ROWS + h
            ws.cell(row=r, column=1, value="N° de Registro")
            ws.cell(row=r, column=2, value="Escribano")
        wb.save(self.path)

        metadatos = self.repo.extraer_metadatos(self.path, fila_datos_inicio=10)
        assert metadatos["siglo"] == "XIX"
        assert metadatos["acervo_num"] == "07"

    def test_detectar_fila_inicio_datos(self):
        _crear_inventario(self.path, data_start_row=12)
        assert self.repo.detectar_fila_inicio_datos(self.path) == 12

    def test_detectar_fila_inicio_datos_con_cabecera_larga(self):
        _crear_inventario(self.path, data_start_row=15)
        assert self.repo.detectar_fila_inicio_datos(self.path) == 15

    def test_bloque_de_ejemplos_y_anotaciones_se_excluye(self):
        _crear_inventario_con_ejemplos(self.path)
        records = self.repo.cargar_registros(
            self.path, fila_datos_inicio=12, fila_inicio=10, fila_fin=500
        )
        assert [r.fila for r in records] == [19, 20, 21]
        assert records[0].registro == "101"
        assert records[0].folios == "001r-002v"

    def test_bloque_de_ejemplos_con_deteccion_automatica(self):
        _crear_inventario_con_ejemplos(self.path)
        fila_detectada = self.repo.detectar_fila_inicio_datos(self.path)
        records = self.repo.cargar_registros(
            self.path, fila_datos_inicio=fila_detectada, fila_inicio=10, fila_fin=500
        )
        assert [r.fila for r in records] == [19, 20, 21]

    def test_titulo_se_mapea_desde_columna_titulo_con_acento(self):
        wb = Workbook()
        ws = wb.active
        for i in range(9):
            ws.cell(row=i + 1, column=1, value=f"Cabecera {i + 1}")
        for h in range(HEADER_ROWS):
            r = 10 + h
            ws.cell(row=r, column=1, value="N° de Registro")
            ws.cell(row=r, column=2, value="Escribano")
            ws.cell(row=r, column=3, value="N° de Folios")
            ws.cell(row=r, column=4, value="N° de Prot")
        ws.cell(row=10, column=8, value="TÍTULO/\n ESCRITURA")
        ws.cell(row=10, column=15, value="RESPONSABLE DE LA DESCRIPCIÓN")
        for i, (reg, esc, fol, prot, tit) in enumerate([
            ("101", "García", "001r-002v", "1", "PODER"),
            ("102", "López", "003r-004v", "1", "VENTA"),
        ]):
            r = 12 + i
            ws.cell(row=r, column=1, value=reg)
            ws.cell(row=r, column=2, value=esc)
            ws.cell(row=r, column=3, value=fol)
            ws.cell(row=r, column=4, value=prot)
            ws.cell(row=r, column=8, value=tit)
            ws.cell(row=r, column=15, value="Heidy Sano")
        wb.save(self.path)

        records = self.repo.cargar_registros(
            self.path, fila_datos_inicio=12, fila_inicio=10, fila_fin=500
        )
        assert [r.titulo for r in records] == ["PODER", "VENTA"]

    def test_detectar_sin_encabezados_devuelve_none(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Sin estructura")
        ws.cell(row=2, column=2, value="solo texto")
        wb.save(self.path)
        assert self.repo.detectar_fila_inicio_datos(self.path) is None
