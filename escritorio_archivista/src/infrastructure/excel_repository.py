"""
Repositorio Excel — implementación del ExcelRepositoryPort.

Carga inventarios archivísticos desde archivos Excel (.xlsx)
usando pandas + openpyxl, con mapeo de columnas y metadatos
configurable a través del archivo maestro (mapa_maestro.json).
"""
import logging
import re
import unicodedata
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook

from src.domain.entities import InventoryRecord
from src.domain.ports.excel_port import ExcelRepositoryPort
from src.domain.value_objects import (
    HEADER_ROWS,
    SKIP_ROW_KEYWORDS,
    ANNOTATION_KEYWORDS,
    EMPTY_VALUES,
)
from src.infrastructure.mapeo_maestro import MapeoMaestro, cargar_mapeo_maestro

logger = logging.getLogger(__name__)


class ExcelRepository(ExcelRepositoryPort):
    """Implementación concreta del repositorio Excel."""

    def __init__(self, mapeo: Optional[MapeoMaestro] = None):
        self._mapeo = mapeo or cargar_mapeo_maestro()

    def cargar_registros(
        self,
        ruta: str,
        fila_datos_inicio: int,
        fila_inicio: int,
        fila_fin: int,
    ) -> List[InventoryRecord]:
        """Carga registros del inventario Excel."""
        logger.info("Cargando registros desde: %s", ruta)

        skiprows = fila_datos_inicio - HEADER_ROWS - 1

        df = pd.read_excel(
            ruta,
            skiprows=skiprows,
            header=list(range(HEADER_ROWS)),
            engine='openpyxl',
        )

        # Aplanar MultiIndex de columnas
        df.columns = [
            ' '.join(str(c) for c in col).strip()
            if isinstance(col, tuple) else str(col)
            for col in df.columns
        ]

        # Mapear columnas
        col_map = self._mapear_columnas(df.columns.tolist())
        logger.info("Columnas mapeadas: %s", col_map)

        records: List[InventoryRecord] = []
        record_counter = 0

        # Filas de anotación/instrucciones (no son datos)
        annot = df.apply(
            lambda row: any(
                kw in self._safe_str(v).lower()
                for v in row for kw in ANNOTATION_KEYWORDS
            ),
            axis=1,
        ).tolist()

        rows = list(df.iterrows())
        for pos, (_, row) in enumerate(rows):
            fila_real = fila_datos_inicio + pos

            # Filtrar por rango de filas
            if fila_real < fila_inicio or fila_real > fila_fin:
                continue

            # Verificar si es fila de anotación/instrucción
            if annot[pos]:
                continue

            # Verificar si es fila de ejemplo (seguida de una anotación)
            if pos + 1 < len(rows) and annot[pos + 1]:
                continue

            # Verificar si es sub-encabezado
            first_cell = self._safe_str(row.iloc[0])
            if any(kw in first_cell.lower() for kw in SKIP_ROW_KEYWORDS):
                continue

            # Extraer valores
            registro = self._get_col(row, col_map, "registro")
            escribano = self._get_col(row, col_map, "escribano")
            folios = self._get_col(row, col_map, "folios")

            # Ignorar filas vacías
            if not registro and not escribano and not folios:
                continue

            record_counter += 1
            protocolo = self._get_col(row, col_map, "protocolo")
            protocolo = self._normalizar_protocolo(protocolo)

            records.append(InventoryRecord(
                id=f"#{record_counter:04d}",
                fila=fila_real,
                registro=registro,
                escribano=escribano,
                protocolo=protocolo,
                folios=folios,
                pg_pdf="",  # Se calcula después con el mapper
                titulo=self._get_col(row, col_map, "titulo"),
                fecha_inicio=self._get_col(row, col_map, "fecha_inicio"),
                fecha_fin=self._get_col(row, col_map, "fecha_fin"),
                interesado1=self._get_col(row, col_map, "interesado1"),
                interesado2=self._get_col(row, col_map, "interesado2"),
                interesado3=self._get_col(row, col_map, "interesado3"),
                data_topica=self._get_col(row, col_map, "data_topica"),
            ))

        logger.info("Cargados %d registros.", len(records))
        return records

    def extraer_metadatos(self, ruta: str, fila_datos_inicio: int) -> dict:
        """Extrae metadatos globales del Excel desde las filas del mapeo maestro."""
        logger.info("Extrayendo metadatos de: %s", ruta)

        metadatos = {
            "filepath": ruta,
            "siglo": "",
            "escribano": "",
            "acervo_num": "",
        }

        try:
            nrows = max(0, fila_datos_inicio - HEADER_ROWS - 1)
            df_header = pd.read_excel(
                ruta, header=None, nrows=nrows, engine='openpyxl'
            )

            # Leer cada metadato desde la (fila, columna) del mapeo maestro
            siglo_ubic = self._mapeo.ubicacion_metadato("siglo")
            escribano_ubic = self._mapeo.ubicacion_metadato("escribano")
            acervo_ubic = self._mapeo.ubicacion_metadato("acervo")

            if siglo_ubic is not None:
                cell = self._celda(df_header, siglo_ubic.fila, siglo_ubic.columna)
                metadatos["siglo"] = self._extraer_siglo_romano(cell)

            if escribano_ubic is not None:
                cell = self._celda(df_header, escribano_ubic.fila, escribano_ubic.columna)
                metadatos["escribano"] = self._extraer_escribano(cell)

            if acervo_ubic is not None:
                cell = self._celda(df_header, acervo_ubic.fila, acervo_ubic.columna)
                metadatos["acervo_num"] = self._extraer_acervo(cell)

        except Exception as e:
            logger.warning("Error extrayendo metadatos: %s", e)

        return metadatos

    @staticmethod
    def _celda(df_header, fila: int, columna: int) -> str:
        """Lee una celda 1-based del DataFrame de cabecera como string."""
        if fila - 1 >= len(df_header):
            return ""
        row = df_header.iloc[fila - 1]
        if columna - 1 >= len(row):
            return ""
        return ExcelRepository._safe_str(row.iloc[columna - 1])

    @staticmethod
    def _extraer_siglo_romano(cell: str) -> str:
        """Busca un siglo romano en el texto de la celda (ej: 'Sección: XIX')."""
        if not cell:
            return ""
        match = re.search(
            r'[:\s]+([IVXLCDM]+)', cell, re.IGNORECASE
        )
        if match:
            return match.group(1).upper()
        return ""

    @staticmethod
    def _extraer_escribano(cell: str) -> str:
        """Obtiene el escribano del texto de la celda.

        Soporta formatos como 'Escribano: Don Pedro' o el nombre directo.
        """
        if not cell:
            return ""
        for sep in (":", "—", "-"):
            if sep in cell:
                return cell.split(sep, 1)[1].strip()
        return cell.strip()

    @staticmethod
    def _extraer_acervo(cell: str) -> str:
        """Extrae el número de acervo del texto (ej: 'Código del fondo: N07')."""
        if not cell:
            return ""
        match = re.search(r'N?(\d+)', cell)
        if match:
            return match.group(1)
        return ""

    def detectar_fila_inicio_datos(self, ruta: str) -> Optional[int]:
        """Detecta la fila 1-based donde empiezan los datos."""
        keywords = ("registro", "escribano", "protocolo", "folios")

        try:
            df = pd.read_excel(
                ruta, header=None, nrows=200, engine='openpyxl'
            )
        except Exception as e:
            logger.warning("No se pudo detectar la fila de datos: %s", e)
            return None

        for idx, row in df.iterrows():
            cells = [self._safe_str(v).lower() for v in row if v is not None]
            non_empty = [c for c in cells if c]
            if len(non_empty) < 3:
                continue
            matches = sum(1 for kw in keywords for c in non_empty if kw in c)
            if matches >= 2:
                return int(idx) + 1 + HEADER_ROWS

        return None

    def guardar_registros(
        self,
        ruta: str,
        fila_datos_inicio: int,
        records: List[InventoryRecord],
    ) -> int:
        """Escribe de vuelta los valores de los registros en el Excel.

        Solo se sobrescriben las columnas localizadas por el mapa maestro
        y los registros con ``fila`` válida. Las celdas cuyo valor en el
        registro está vacío no se tocan, para no borrar contenido que no
        fue parseado.
        """
        try:
            wb = load_workbook(ruta)
        except Exception as e:
            logger.error("No se pudo abrir el Excel para guardar: %s", e)
            raise

        ws = wb.worksheets[0]
        columnas = self._indices_columnas(ws, fila_datos_inicio)

        celdas = 0
        for record in records:
            fila = record.fila
            if not fila:
                continue
            for campo, col in columnas.items():
                valor = getattr(record, campo, "")
                if valor in (None, ""):
                    continue
                ws.cell(row=fila, column=col).value = valor
                celdas += 1

        wb.save(ruta)
        logger.info("Guardados %d valores en el Excel %s.", celdas, ruta)
        return celdas

    def _indices_columnas(self, ws, fila_datos_inicio: int) -> dict:
        """Localiza la columna (1-based) de cada campo en las filas de encabezado.

        Recrea el texto combinado de cada columna a partir de las
        ``HEADER_ROWS`` filas previas a los datos, igual que hace
        ``_mapear_columnas`` con el MultiIndex de pandas.
        """
        ncols = ws.max_column or 0
        encabezado = []
        for r in range(fila_datos_inicio - HEADER_ROWS, fila_datos_inicio):
            fila_texto = [
                self._safe_str(ws.cell(row=r, column=c).value)
                for c in range(1, ncols + 1)
            ]
            encabezado.append(fila_texto)

        indices = {}
        for campo, mapeo in self._mapeo.columnas.items():
            idx = None
            for c in range(ncols):
                texto = self._sin_acentos(
                    " ".join(fila[c] for fila in encabezado).lower()
                )
                if any(
                    self._sin_acentos(a.lower()) in texto
                    for a in mapeo.aliases
                ):
                    idx = c
                    break
            if idx is None and mapeo.indice_fallback is not None:
                idx = mapeo.indice_fallback
            if idx is not None:
                indices[campo] = idx + 1

        return indices

    def _mapear_columnas(self, columnas: List[str]) -> dict:
        """Mapea nombres de columnas del Excel a campos del modelo.

        Usa el mapeo maestro (aliases + índice de respaldo). Si no hay un
        campo mapeado, se lo omite del resultado.
        """
        col_map = {}
        columnas_lower = [self._sin_acentos(c.lower()) for c in columnas]

        for campo, mapeo in self._mapeo.columnas.items():
            for alias in mapeo.aliases:
                alias_norm = self._sin_acentos(alias)
                for i, col_lower in enumerate(columnas_lower):
                    if alias_norm in col_lower:
                        col_map[campo] = i
                        break
                if campo in col_map:
                    break

            # Fallback por índice
            if campo not in col_map and mapeo.indice_fallback is not None:
                if mapeo.indice_fallback < len(columnas):
                    col_map[campo] = mapeo.indice_fallback

        return col_map

    @staticmethod
    def _sin_acentos(texto: str) -> str:
        """Elimina acentos/diacríticos para comparación insensible a acentos."""
        return ''.join(
            c for c in unicodedata.normalize('NFD', texto)
            if unicodedata.category(c) != 'Mn'
        )

    def _get_col(self, row, col_map: dict, campo: str) -> str:
        """Obtiene el valor de una columna mapeada como string limpio."""
        if campo not in col_map:
            return ""
        idx = col_map[campo]
        if idx >= len(row):
            return ""
        val = row.iloc[idx]
        return self._safe_str(val)

    @staticmethod
    def _safe_str(val) -> str:
        """Convierte un valor a string, tratando NaN/None como vacío."""
        if val is None:
            return ""
        s = str(val).strip()
        if s.lower() in EMPTY_VALUES:
            return ""
        return s

    @staticmethod
    def _normalizar_protocolo(val: str) -> str:
        """Normaliza protocolo: si es float entero (ej: '3.0'), convierte a '3'."""
        if not val:
            return ""
        try:
            f = float(val)
            if f == int(f):
                return str(int(f))
        except (ValueError, TypeError):
            pass
        return val
